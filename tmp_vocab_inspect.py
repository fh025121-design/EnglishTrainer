from openpyxl import load_workbook
p = r'c:\Users\hiroshiinoue\Downloads\GitHub\EnglishTrainer\mobile\vocabulary-master 0822.xlsx'
wb = load_workbook(p, data_only=True)
print('SHEETS', wb.sheetnames)
for ws in wb.worksheets:
    rows = []
    for r in ws.iter_rows(values_only=True):
        values = [None if v is None else str(v).strip() for v in r]
        if not any(v is not None for v in values):
            continue
        if values[1] in ('レベル', 'Level'):
            continue
        if len(values) >= 8 and values[1] not in ('', None):
            rows.append(values)
    print(ws.title, len(rows), 'rows')
    print(rows[:3])
