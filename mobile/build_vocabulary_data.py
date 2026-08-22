import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / 'vocabulary-master 0822.xlsx'
OUTPUT_PATH = ROOT / 'vocabulary-data.js'

GRADE_ORDER = ['5', '4', '3']
TARGET_SHEETS = {'5級': '5', '4級': '4', '3級': '3'}


def normalize_text(value):
    if value is None:
        return ''
    return str(value).strip()


def slugify(value):
    slug = re.sub(r'[^a-z0-9]+', '-', value.lower())
    slug = slug.strip('-')
    return slug or 'item'


def build_id(grade, word, part_of_speech, index):
    core = f'{grade}-{slugify(word)}-{slugify(part_of_speech)}-{index:03d}'
    return f'vocab-{core}'


def row_to_entry(row, grade, index):
    cells = [normalize_text(v) for v in row]
    if len(cells) < 8:
        return None
    level, word, part_of_speech, meaning, accent, example, example_ja = cells[1:8]
    if not level or not word:
        return None
    if level not in ('5級', '4級', '3級'):
        return None
    if any(ch.isspace() for ch in word):
        return None
    if not part_of_speech or not meaning or not accent:
        return None
    accent_focus = ''
    if accent:
        uppercase_matches = re.findall(r'[A-Z]+', accent)
        if uppercase_matches:
            accent_focus = uppercase_matches[0]
        else:
            accent_focus = accent
    entry = {
        'id': build_id(grade, word, part_of_speech, index),
        'grade': int(grade),
        'word': word,
        'partOfSpeech': part_of_speech,
        'meaning': meaning,
        'accent': accent,
        'accentFocus': accent_focus,
        'exampleSentence': example,
        'exampleTranslation': example_ja,
    }
    return entry


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f'Excel file not found: {XLSX_PATH}')

    wb = load_workbook(XLSX_PATH, data_only=True)
    merged = []
    for sheet_name, grade in TARGET_SHEETS.items():
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if row[1] is None:
                continue
            first = normalize_text(row[1])
            if first in ('レベル', 'Level'):
                continue
            if not any(v is not None for v in row):
                continue
            entry = row_to_entry(row, grade, len(merged) + 1)
            if entry:
                merged.append(entry)

    deduped = []
    seen = set()
    for entry in merged:
        sig = (entry['word'], entry['partOfSpeech'], entry['meaning'])
        if sig in seen:
            continue
        seen.add(sig)
        deduped.append(entry)

    deduped.sort(key=lambda item: (GRADE_ORDER.index(str(item['grade'])), item['word'], item['partOfSpeech']))

    lines = []
    lines.append('(function () {')
    lines.append('  const entries = [')
    for entry in deduped:
        lines.append('    {')
        lines.append(f"      id: \"{entry['id']}\",")
        lines.append(f"      grade: {entry['grade']},")
        lines.append(f"      word: \"{entry['word']}\",")
        lines.append(f"      partOfSpeech: \"{entry['partOfSpeech']}\",")
        lines.append(f"      meaning: \"{entry['meaning']}\",")
        lines.append(f"      accent: \"{entry['accent']}\",")
        lines.append(f"      accentFocus: \"{entry['accentFocus']}\",")
        lines.append(f"      exampleSentence: \"{entry['exampleSentence']}\",")
        lines.append(f"      exampleTranslation: \"{entry['exampleTranslation']}\",")
        lines.append('    },')
    lines.append('  ];')
    lines.append('  window.MOBILE_VOCABULARY_REAL_WORD_BANK = entries;')
    lines.append('})();')

    OUTPUT_PATH.write_text('\n'.join(lines) + '\n', encoding='utf-8')

    print(f'Generated {len(deduped)} entries from {XLSX_PATH.name}')
    counts = {'5': 0, '4': 0, '3': 0}
    for entry in deduped:
        counts[str(entry['grade'])] += 1
    print('Counts:', counts)


if __name__ == '__main__':
    main()
